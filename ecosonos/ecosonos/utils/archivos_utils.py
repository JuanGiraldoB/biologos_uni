import pandas as pd
import shutil
import os
from datetime import datetime
import csv
import openpyxl as xl
import pathlib


def mover_archivos_segun_tipo(carpeta_raiz, carpeta_destino, tipo):
    ruta_csv = f'{carpeta_raiz}/resultado_preproceso.csv'
    csv_file = pd.read_csv(ruta_csv)

    print(csv_file.head())

    for _, row in csv_file.iterrows():
        ruta_archivo = row['path_FI']
        lluvia = row['rain_FI']

        if lluvia == tipo:
            shutil.move(ruta_archivo, carpeta_destino)


def obtener_archivos_wav(carpetas):
    archivos = []
    formatos = ['wav', 'WAV', 'mp3']
    nombres_base = []

    for carpeta in carpetas:
        for archivo in os.listdir(carpeta):
            file_name = os.path.join(
                carpeta, archivo).replace(os.path.sep, '/')

            # Verifica que sea un archivo y que la extension corresponda con las de la variable formatos
            if os.path.isfile(file_name):
                file_extension = file_name.split(".")[1]
                if file_extension in formatos:
                    nombre_base = os.path.basename(file_name)
                    nombres_base.append(nombre_base)
                    archivos.append(file_name)

    return archivos, nombres_base


def reemplazar_caracter(archivos, caracter, reemplazo):
    for i in range(len(archivos)):
        archivos[i] = archivos[i].replace(caracter, reemplazo)


def obtener_fecha(archivo):

    if "." in archivo:
        archivo = pathlib.Path(archivo).stem

    archivo_partes = archivo.split('_')
    fecha = archivo_partes[-2]
    hora = archivo_partes[-1]

    yy = int(fecha[0:4])
    mm = int(fecha[4:6])
    dd = int(fecha[6:8])

    hh = int(hora[0:2])
    min = int(hora[2:4])
    ss = int(hora[4:6])

    date = datetime(year=yy, month=mm, day=dd, hour=hh, minute=min, second=ss,
                    microsecond=0)

    return date


def crear_csv(carpeta_raiz, nombre_carpeta):
    nombre_csv = f'etiquetas_{nombre_carpeta}.csv'
    csv_ruta = os.path.join(carpeta_raiz, nombre_csv)

    if not os.path.exists(csv_ruta):
        with open(csv_ruta, mode='w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["grabacion", "etiqueta", "t_min",
                            "t_max", "f_min", "f_max"])

    return csv_ruta


def agregar_fila_csv(csv_ruta, nombre_grabacion, etiqueta, x0, x1, y0, y1):
    with open(csv_ruta, "a", newline='') as file:
        writer = csv.writer(file)

        fila = [nombre_grabacion, etiqueta, x0, x1, y0, y1]
        writer.writerow(fila)


# def crear_xlsx(carpeta_raiz):
#     base = os.path.basename(carpeta_raiz)
#     nombre = f'etiquetas-{base}.xlsx'
#     xlsx_ruta = os.path.join(carpeta_raiz, nombre).replace('\\', '/')

#     print(xlsx_ruta)
#     if not os.path.exists(xlsx_ruta):
#         wb = xl.Workbook()
#         wb.save(xlsx_ruta)
#         wb.close()

#     return xlsx_ruta


# def crear_hoja_xlsx(xlsx_ruta, nombre_archivo):
#     wb = xl.load_workbook(filename=xlsx_ruta)  # wb = workbook
#     nombre_hojas = wb.sheetnames

#     if nombre_archivo in nombre_hojas:
#         ws = wb[nombre_archivo]
#     else:
#         ws = wb.create_sheet(title=nombre_archivo)
#         ws.append(["etiqueta,t_min,t_max,f_min,f_max"])

#     if "Sheet" in nombre_hojas:
#         del wb["Sheet"]

#     wb.save(xlsx_ruta)
#     wb.close()


# def agregar_fila_csv(csv_ruta, nombre_archivo, etiqueta, x0, x1, y0, y1):
#     wb = xl.load_workbook(filename=csv_ruta)
#     ws = wb[nombre_archivo]

#     fila = [f"{etiqueta},{x0},{x1},{y0},{y1}"]
#     ws.append(fila)

#     wb.save(csv_ruta)
#     wb.close()


def selecciono_archivo(archivo):
    """
        Verifica si fue seleccionada un archivo
    """
    return not archivo
